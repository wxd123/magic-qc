# interface/facial/export.py
"""
结果导出模块 - 负责所有结果保存相关的函数
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from rich.console import Console

from magic_qc.interface.facial.display import get_status_and_color

console = Console()


# ============================================================
# 单张结果保存
# ============================================================

def save_single_result(result: Dict[str, Any], output_path: str) -> None:
    """
    保存单张图片检测结果
    
    Args:
        result: 检测结果字典
        output_path: 输出文件路径
    """
    output_file = Path(output_path)
    
    if output_file.suffix.lower() == '.csv':
        _save_single_result_as_csv(result, output_file)
    else:
        _save_single_result_as_json(result, output_file)
    
    console.print(f"\n💾 结果已保存至: {output_path}")


def _save_single_result_as_csv(result: Dict[str, Any], output_file: Path) -> None:
    """保存单张结果为CSV格式"""
    score = result.get('score', 0)
    status, _ = get_status_and_color(score)
    
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['指标', '数值'])
        writer.writerow(['文件名', result.get('filename', 'N/A')])
        writer.writerow(['是否通过', '通过' if result.get('passed') else '不通过'])
        writer.writerow(['建模适用性', status])
        writer.writerow(['综合评分', round(score, 1)])
        
        # 新增：独立指标
        writer.writerow(['对称性分数', result.get('overall_symmetry', 0)])
        writer.writerow(['亮度比', result.get('lr_ratio', 1.0)])
        writer.writerow(['清晰度', result.get('clarity', 0)])
        writer.writerow(['对比度', result.get('contrast', 0)])
        writer.writerow(['边缘密度', result.get('edge_density', 0)])
        writer.writerow(['噪声水平', result.get('noise_level', 0)])
        writer.writerow(['质量评分', result.get('quality_score', 0)])
        
        # 保留原有详细数据（兼容）
        symmetry_data = result.get('symmetry', {})
        quality_data = result.get('quality', {})
        quality_metrics = quality_data.get('metrics', {})
        
        writer.writerow(['对称性等级', symmetry_data.get('symmetry_level', 'N/A')])
        writer.writerow(['真实性评分', symmetry_data.get('authenticity_score', 0)])
        writer.writerow(['是否真实人脸', '真实人脸' if symmetry_data.get('is_realistic') else 'AI生成'])
        writer.writerow(['质量等级', quality_data.get('quality_level', 'N/A')])
        writer.writerow(['建议', quality_data.get('recommendation', 'N/A')])


def _save_single_result_as_json(result: Dict[str, Any], output_file: Path) -> None:
    """保存单张结果为JSON格式"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)


# ============================================================
# 批量结果保存
# ============================================================

def save_batch_results(
    results: List[Dict[str, Any]], 
    output_path: str, 
    sort_by_score: bool = True, 
    descending: bool = True
) -> None:
    """
    保存批量检测结果
    
    Args:
        results: 检测结果列表
        output_path: 输出文件路径
        sort_by_score: 是否按评分排序
        descending: 是否降序
    """
    # 复制结果并排序
    sorted_results = results.copy()
    if sort_by_score:
        sorted_results.sort(key=lambda x: x.get('score', 0), reverse=descending)
    
    output_file = Path(output_path)
    
    if output_file.suffix.lower() == '.csv':
        _save_batch_results_as_csv(sorted_results, output_file)
    else:
        _save_batch_results_as_json(sorted_results, output_file)
    
    console.print(f"\n💾 结果已保存至: {output_path}（按综合评分{'降序' if descending else '升序'}排列）")


def _save_batch_results_as_csv(sorted_results: List[Dict[str, Any]], output_file: Path) -> None:
    """保存批量结果为CSV格式"""
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        # 表头：新增指标 + 原有指标
        writer.writerow([
            '排名', '文件名',
            # 新增指标
            '对称性分数', '亮度比', '清晰度', '对比度', '噪声水平', '边缘密度',
            '综合评分', '建模适用性', '质量评分',
            # 原有指标
            '对称性等级', '真实性评分',  '建议'
        ])
        
        for rank, r in enumerate(sorted_results, 1):
            score = r.get('score', 0)
            status, _ = get_status_and_color(score)
            
            # 获取原有数据
            symmetry_data = r.get('symmetry', {})
            quality_data = r.get('quality', {})
            
            writer.writerow([
                rank,
                r.get('filename', 'N/A'),
                # 新增指标
                r.get('overall_symmetry', 0),
                r.get('lr_ratio', 1.0),
                r.get('clarity', 0),
                r.get('contrast', 0),
                r.get('noise_level', 0),
                r.get('edge_density', 0),
                round(score, 1),
                status,
                r.get('quality_score', 0),
                # 原有指标
                symmetry_data.get('symmetry_level', 'N/A'),
                symmetry_data.get('authenticity_score', 0),                
                quality_data.get('recommendation', 'N/A')
            ])


def _save_batch_results_as_json(sorted_results: List[Dict[str, Any]], output_file: Path) -> None:
    """保存批量结果为JSON格式"""
    output_data = {
        "total": len(sorted_results),
        "sort_by": "score",
        "sort_order": "descending",
        "results": sorted_results
    }
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)


# ============================================================
# 对称性结果保存
# ============================================================

def save_symmetry_batch_results(
    results: List[Dict[str, Any]], 
    output_path: str, 
    sort_by_score: bool = True, 
    descending: bool = True
) -> None:
    """
    保存对称性批量分析结果
    
    Args:
        results: 对称性分析结果列表
        output_path: 输出文件路径
        sort_by_score: 是否按对称性分数排序
        descending: 是否降序
    """
    # 复制结果并排序
    sorted_results = results.copy()
    if sort_by_score:
        sorted_results.sort(key=lambda x: x.get('overall_symmetry', 0), reverse=descending)
    
    output_file = Path(output_path)
    
    if output_file.suffix.lower() == '.csv':
        _save_symmetry_results_as_csv(sorted_results, output_file)
    else:
        _save_symmetry_results_as_json(sorted_results, output_file)
    
    console.print(f"\n💾 结果已保存至: {output_path}（按对称性分数{'降序' if descending else '升序'}排列）")


def _save_symmetry_results_as_csv(sorted_results: List[Dict[str, Any]], output_file: Path) -> None:
    """保存对称性结果为CSV格式"""
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['排名', '文件名', '对称性分数', '对称性等级', '亮度比', '真实性评分', '是否真实'])
        for rank, r in enumerate(sorted_results, 1):
            writer.writerow([
                rank,
                r.get('filename', 'N/A'),
                r.get('overall_symmetry', 0),
                r.get('symmetry_level', 'N/A'),
                r.get('lr_ratio', 1.0),
                r.get('authenticity_score', 0),
                '真实' if r.get('is_realistic') else 'AI生成'
            ])


def _save_symmetry_results_as_json(sorted_results: List[Dict[str, Any]], output_file: Path) -> None:
    """保存对称性结果为JSON格式"""
    output_data = {
        "total": len(sorted_results),
        "sort_by": "symmetry_score",
        "sort_order": "descending",
        "results": sorted_results
    }
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)


# ============================================================
# 智能输出路径处理
# ============================================================

def resolve_output_path(
    input_path: str, 
    output_arg: Optional[str], 
    default_filename: str = "report.csv"
) -> Optional[Path]:
    """
    智能解析输出路径
    
    Args:
        input_path: 输入文件/目录路径
        output_arg: 命令行传入的 -o 参数
        default_filename: 默认文件名
    
    Returns:
        输出路径，如果不需要保存则返回 None
    
    使用规则:
        - output_arg 为 None: 不保存
        - output_arg 为空字符串或 ".": 在输入目录生成默认文件名
        - output_arg 以 '/' 结尾: 作为目录，生成默认文件名
        - output_arg 是已存在的目录: 在该目录生成默认文件名
        - 其他: 作为完整文件路径
    """
    if output_arg is None:
        return None
    
    input_path_obj = Path(input_path)
    output_str = str(output_arg).strip()
    
    # 情况1: -o 不带参数（typer 中可能传空字符串或特殊值）
    if output_str == "" or output_str == ".":
        if input_path_obj.is_dir():
            return input_path_obj / default_filename
        else:
            return input_path_obj.parent / f"{input_path_obj.stem}_{default_filename}"
    
    output_path = Path(output_str)
    
    # 情况2: 以 / 结尾，作为目录
    if output_str.endswith('/'):
        output_path.mkdir(parents=True, exist_ok=True)
        if input_path_obj.is_dir():
            return output_path / default_filename
        else:
            return output_path / f"{input_path_obj.stem}_{default_filename}"
    
    # 情况3: 是已存在的目录
    if output_path.exists() and output_path.is_dir():
        if input_path_obj.is_dir():
            return output_path / default_filename
        else:
            return output_path / f"{input_path_obj.stem}_{default_filename}"
    
    # 情况4: 作为文件路径
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


# ============================================================
# 导出格式转换
# ============================================================

def export_to_excel(results: List[Dict[str, Any]], output_path: str) -> None:
    """
    导出为Excel格式（需要安装 openpyxl）
    
    Args:
        results: 检测结果列表
        output_path: 输出文件路径
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        console.print("[red]❌ 请安装 openpyxl: pip install openpyxl[/red]")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "三维建模质控报告"
    
    # 更新表头：与display表格保持一致
    headers = ['排名', '文件名', '对称性', '亮度比', '清晰度', '对比度', '噪点', '综合评分', '建模适用性', '质量评分']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 数据
    for row, r in enumerate(sorted_results, 2):
        score = r.get('score', 0)
        status, _ = get_status_and_color(score)
        
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=r.get('filename', 'N/A'))
        ws.cell(row=row, column=3, value=r.get('overall_symmetry', 0))
        ws.cell(row=row, column=4, value=r.get('lr_ratio', 1.0))
        ws.cell(row=row, column=5, value=r.get('clarity', 0))
        ws.cell(row=row, column=6, value=r.get('contrast', 0))
        ws.cell(row=row, column=7, value=r.get('noise_level', 0))
        ws.cell(row=row, column=8, value=round(score, 1))
        ws.cell(row=row, column=9, value=status)
        ws.cell(row=row, column=10, value=r.get('quality_score', 0))
    
    wb.save(output_path)
    console.print(f"\n💾 Excel报告已保存至: {output_path}")